"""
gui.py  —  CustomTkinter UI for LinkedIn Job Scraper
Drop this file alongside cli.py, config.py, scraper.py, etc.
Run:  python gui.py
Requires: pip install customtkinter playwright openpyxl
"""

import asyncio
import json
import threading
import webbrowser
from datetime import datetime
from tkinter import messagebox

import customtkinter as ctk

from .config import SearchConfig, JobListing
from .scraper import LinkedInScraper
from .url_builder import build_url

# ─────────────────────────────────────────────
#  Theme palettes
# ─────────────────────────────────────────────
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("dark-blue")

PALETTE_DARK_KHAKI_ORANGE = {
    "bg":          "#081416",   # midnight teal-black
    "panel":       "#102225",   # deep teal panel
    "card":        "#182d30",   # dark teal card surface
    "card_hover":  "#21393d",   # lifted teal hover state
    "border":      "#3f5e62",   # muted teal border
    "accent":      "#e39a63",   # apricot primary
    "accent2":     "#f2c08f",   # soft apricot highlight
    "warn":        "#c75f4a",   # coral-clay warning
    "text":        "#f3eadf",   # warm cream text
    "muted":       "#94a7a7",   # muted teal-gray text
    "tag_bg":      "#2f241d",   # dark apricot-teal tag background
    "tag_text":    "#efa76f",   # apricot tag text
}

PALETTE_SOFT_SLATE_KHAKI_ORANGE = {
    "bg":          "#eef1ef",
    "panel":       "#e1e6e3",
    "card":        "#f8faf7",
    "card_hover":  "#edf1ed",
    "border":      "#c9d0cc",
    "accent":      "#d9772f",
    "accent2":     "#b79a58",
    "warn":        "#b85c38",
    "text":        "#243033",
    "muted":       "#6f7a76",
    "tag_bg":      "#f2ddc7",
    "tag_text":    "#b85f28",
    "_label":      "Soft Slate",
    "_mode":       "light",
}

PALETTE_MINT_LATTE = {
    "bg":          "#edf7f1",
    "panel":       "#e0efe7",
    "card":        "#fbf7ef",
    "card_hover":  "#f0eadf",
    "border":      "#c9d8cf",
    "accent":      "#4fae83",
    "accent2":     "#c8a77c",
    "warn":        "#b86b4b",
    "text":        "#24322b",
    "muted":       "#718176",
    "tag_bg":      "#d9efe5",
    "tag_text":    "#2f8f68",
    "_label":      "Mint Latte",
    "_mode":       "light",
}

PALETTE_DARK_SLATE_KHAKI = {
    "bg":          "#1f2426",
    "panel":       "#2a3032",
    "card":        "#343a3c",
    "card_hover":  "#3f4648",
    "border":      "#6f674e",
    "accent":      "#d9772f",
    "accent2":     "#c6a15b",
    "warn":        "#b85c38",
    "text":        "#efe6d2",
    "muted":       "#9a927d",
    "tag_bg":      "#3a3328",
    "tag_text":    "#e28a3f",
    "_label":      "Dark Slate",
    "_mode":       "dark",
}

ALL_PALETTES = [
    PALETTE_DARK_KHAKI_ORANGE,
    PALETTE_SOFT_SLATE_KHAKI_ORANGE,
    PALETTE_MINT_LATTE,
    PALETTE_DARK_SLATE_KHAKI,
]
 
# Active palette (mutable reference)
PALETTE = dict(PALETTE_DARK_KHAKI_ORANGE)
 
 
FONT_TITLE   = ("Bebas Neue", 21, "bold")
FONT_LABEL   = ("Bebas Neue", 11, "bold")
FONT_BODY    = ("Bebas Neue", 11)
FONT_SMALL   = ("Bebas Neue", 10)
FONT_MONO    = ("Bebas Neue", 10)
FONT_SECTION = ("Bebas Neue", 8, "bold")
 
# ─────────────────────────────────────────────
#  Option maps  (label → key)
# ─────────────────────────────────────────────
DATE_OPTIONS = {
    "Any time": "any", "Last hour": "1h", "Last 6 hours": "6h",
    "Last 24 hours": "24h", "Last 3 days": "3d",
    "Last week": "week", "Last month": "month",
}
EXPERIENCE_OPTIONS = {
    "Internship": "internship", "Entry level": "entry",
    "Associate": "associate", "Mid-Senior": "mid",
    "Director": "director", "Executive": "executive",
}
JOB_TYPE_OPTIONS = {
    "Full-time": "fulltime", "Part-time": "parttime",
    "Contract": "contract", "Temporary": "temporary",
    "Internship": "internship",
}
WORK_TYPE_OPTIONS = {
    "On-site": "onsite", "Remote": "remote", "Hybrid": "hybrid",
}
SORT_OPTIONS = {"Most recent": "recent", "Most relevant": "relevant"}
 
 
# ─────────────────────────────────────────────
#  Reusable widgets
# ─────────────────────────────────────────────
 
class SectionLabel(ctk.CTkLabel):
    """Small all-caps muted section heading."""
    def __init__(self, parent, text, **kw):
        super().__init__(
            parent, text=text.upper(),
            font=FONT_SECTION,
            text_color=PALETTE["muted"],
            **kw,
        )
 
 
class MultiCheckGroup(ctk.CTkFrame):
    """
    Vertical stack of checkboxes (one per row) for multi-select filters.
    """
    def __init__(self, parent, options: dict[str, str], **kw):
        super().__init__(parent, fg_color="transparent", **kw)
        self._vars: dict[str, ctk.BooleanVar] = {}
        for label, key in options.items():
            var = ctk.BooleanVar()
            cb = ctk.CTkCheckBox(
                self, text=label, variable=var,
                font=FONT_SMALL,
                text_color=PALETTE["text"],
                fg_color=PALETTE["accent"],
                hover_color=self._accent_hover(),
                border_color=PALETTE["border"],
                checkmark_color="#ffffff",
                width=0,
            )
            cb.pack(anchor="w", pady=2)
            self._vars[key] = var
 
    def _accent_hover(self):
        acc = PALETTE["accent"]
        # darken slightly
        r = max(0, int(acc[1:3], 16) - 30)
        g = max(0, int(acc[3:5], 16) - 30)
        b = max(0, int(acc[5:7], 16) - 30)
        return f"#{r:02x}{g:02x}{b:02x}"
 
    def selected(self) -> list[str]:
        return [k for k, v in self._vars.items() if v.get()]
 
 
def _divider(parent):
    ctk.CTkFrame(parent, fg_color=PALETTE["border"], height=1).pack(
        fill="x", pady=10,
    )
 
 
# ─────────────────────────────────────────────
#  Collapsible section widget
# ─────────────────────────────────────────────
 
class CollapsibleSection(ctk.CTkFrame):
    """
    A section with a toggle button header.
    Content is hidden by default; clicking the header reveals it.
    """
    def __init__(self, parent, title: str, initially_open: bool = False, **kw):
        super().__init__(parent, fg_color="transparent", **kw)
        self._open = initially_open
 
        # ── Toggle button ────────────────────
        self._btn = ctk.CTkButton(
            self,
            text=self._btn_text(title),
            font=("Bebas Neue", 10, "bold"),
            text_color=PALETTE["accent"],
            fg_color=PALETTE["card"],
            hover_color=PALETTE["card_hover"],
            border_width=1,
            border_color=PALETTE["border"],
            corner_radius=5,
            height=30,
            anchor="w",
            command=self._toggle,
        )
        self._btn.pack(fill="x", pady=(0, 4))
        self._title = title
 
        # ── Content frame ────────────────────
        self._content = ctk.CTkFrame(self, fg_color="transparent")
        if initially_open:
            self._content.pack(fill="x")
 
    def _btn_text(self, title):
        arrow = "▼" if self._open else "▶"
        return f"  {arrow}  {title.upper()}"
 
    def _toggle(self):
        self._open = not self._open
        self._btn.configure(text=self._btn_text(self._title))
        if self._open:
            self._content.pack(fill="x", pady=(4, 0))
        else:
            self._content.pack_forget()
 
    @property
    def content(self):
        """Place child widgets inside this frame."""
        return self._content
 
 
# ─────────────────────────────────────────────
#  Theme swatch button
# ─────────────────────────────────────────────
 
class ThemeSwatch(ctk.CTkFrame):
    """Small clickable color swatch for palette selection."""
 
    def __init__(self, parent, palette: dict, on_select, **kw):
        super().__init__(
            parent,
            fg_color=palette["bg"],
            corner_radius=6,
            border_width=2,
            border_color=palette["accent"],
            cursor="hand2",
            width=40, height=40,
            **kw,
        )
        self._palette = palette
        self.pack_propagate(False)
 
        # Mini accent stripe
        ctk.CTkFrame(
            self,
            fg_color=palette["accent"],
            corner_radius=0,
            height=8,
        ).pack(side="bottom", fill="x")
 
        # Tooltip label
        ctk.CTkLabel(
            self,
            text=palette.get("_label", "")[:6],
            font=("Bebas Neue", 7, "bold"),
            text_color=palette["muted"],
        ).pack(expand=True)
 
        self.bind("<Button-1>", lambda e: on_select(palette))
        for child in self.winfo_children():
            child.bind("<Button-1>", lambda e: on_select(palette))
 
    def set_active(self, active: bool):
        self.configure(border_width=3 if active else 2)
 
 
# ─────────────────────────────────────────────
#  Job detail view
# ─────────────────────────────────────────────
 
class JobDetailView(ctk.CTkFrame):
    """Full-screen detail panel shown when a card is clicked."""
 
    def __init__(self, parent, job: JobListing, on_back, **kw):
        super().__init__(parent, fg_color=PALETTE["bg"], corner_radius=0, **kw)
        self._build(job, on_back)
 
    def _build(self, job: JobListing, on_back):
        topbar = ctk.CTkFrame(self, fg_color=PALETTE["panel"], corner_radius=0, height=52)
        topbar.pack(fill="x")
        topbar.pack_propagate(False)
 
        ctk.CTkButton(
            topbar, text="← Back",
            font=FONT_LABEL,
            fg_color="transparent",
            hover_color=PALETTE["card_hover"],
            text_color=PALETTE["accent"],
            border_width=1,
            border_color=PALETTE["border"],
            corner_radius=5,
            height=32, width=90,
            command=on_back,
        ).pack(side="left", padx=16, pady=10)
 
        ctk.CTkLabel(
            topbar, text="JOB DETAILS",
            font=("Bebas Neue", 11, "bold"),
            text_color=PALETTE["muted"],
        ).pack(side="left", padx=4, pady=10)
 
        if job.url:
            ctk.CTkButton(
                topbar, text="Open on LinkedIn ↗",
                font=FONT_SMALL,
                fg_color=PALETTE["accent"],
                hover_color=self._accent_hover(),
                text_color="#fff",
                corner_radius=5,
                height=32,
                command=lambda: webbrowser.open(job.url),
            ).pack(side="right", padx=16, pady=10)
 
        body = ctk.CTkScrollableFrame(
            self, fg_color="transparent",
            scrollbar_button_color=PALETTE["border"],
        )
        body.pack(fill="both", expand=True, padx=32, pady=24)
        self._body = body   # expose for scroll binding after population
 
        ctk.CTkLabel(
            body, text=job.title or "—",
            font=("Bebas Neue", 20, "bold"),
            text_color=PALETTE["text"],
            anchor="w", wraplength=800,
        ).pack(anchor="w", pady=(0, 6))
 
        tags_row = ctk.CTkFrame(body, fg_color="transparent")
        tags_row.pack(anchor="w", pady=(0, 16))
 
        def _tag(text, color=PALETTE["tag_text"]):
            ctk.CTkLabel(
                tags_row, text=text,
                font=FONT_SMALL,
                text_color=color,
                fg_color=PALETTE["tag_bg"],
                corner_radius=4,
                padx=8, pady=3,
            ).pack(side="left", padx=(0, 8))
 
        if job.easy_apply:
            _tag("⚡ Easy Apply", PALETTE["accent2"])
 
        grid = ctk.CTkFrame(
            body, fg_color=PALETTE["panel"],
            corner_radius=8, border_width=1,
            border_color=PALETTE["border"],
        )
        grid.pack(fill="x", pady=(0, 20))
 
        def _row(label, value, value_color=PALETTE["text"]):
            if not value:
                return
            row = ctk.CTkFrame(grid, fg_color="transparent")
            row.pack(fill="x", padx=18, pady=6)
            ctk.CTkLabel(
                row, text=label,
                font=("Bebas Neue", 10, "bold"),
                text_color=PALETTE["muted"],
                width=140, anchor="w",
            ).pack(side="left")
            ctk.CTkLabel(
                row, text=str(value),
                font=FONT_BODY,
                text_color=value_color,
                anchor="w", wraplength=600,
            ).pack(side="left", fill="x", expand=True)
 
        _row("🏢  Company", job.company)
        _row("📍  Location", job.location)
        _row("📅  Date Posted", job.date_posted)
        _row("👥  Applicants", job.num_applicants)
        salary = getattr(job, "salary", None)
        if salary:
            _row("💰  Salary", salary, PALETTE["accent2"])
        _row("🔗  URL", job.url, PALETTE["accent"])
 
        if job.description:
            _divider(body)
            ctk.CTkLabel(
                body, text="DESCRIPTION",
                font=FONT_SECTION, text_color=PALETTE["muted"],
                anchor="w",
            ).pack(anchor="w", pady=(0, 10))
 
            ctk.CTkLabel(
                body,
                text=job.description,
                font=FONT_BODY,
                text_color=PALETTE["text"],
                anchor="nw",
                justify="left",
                wraplength=820,
            ).pack(anchor="w", fill="x")
 
    def _accent_hover(self):
        acc = PALETTE["accent"]
        r = max(0, int(acc[1:3], 16) - 30)
        g = max(0, int(acc[3:5], 16) - 30)
        b = max(0, int(acc[5:7], 16) - 30)
        return f"#{r:02x}{g:02x}{b:02x}"
 
 
# ─────────────────────────────────────────────
#  Job card (list view)
# ─────────────────────────────────────────────
 
class JobCard(ctk.CTkFrame):
    """Single job result card — click to open detail view."""
 
    def __init__(self, parent, job: JobListing, index: int, on_click, **kw):
        super().__init__(
            parent,
            fg_color=PALETTE["card"],
            corner_radius=8,
            border_width=1,
            border_color=PALETTE["border"],
            cursor="hand2",
            **kw,
        )
        self.job = job
        self._on_click = on_click
        self._build(index)
 
        self.bind("<Enter>", lambda e: self.configure(fg_color=PALETTE["card_hover"]))
        self.bind("<Leave>", lambda e: self.configure(fg_color=PALETTE["card"]))
 
        # Bind click on self and every descendant — no add="+", so no stacking
        def _bind_click(w):
            w.bind("<Button-1>", lambda e, j=job: on_click(j))
            for child in w.winfo_children():
                _bind_click(child)
 
        _bind_click(self)
 
    def _build(self, idx: int):
        top = ctk.CTkFrame(self, fg_color="transparent")
        top.pack(fill="x", padx=14, pady=(10, 4))
 
        ctk.CTkLabel(
            top, text=f"{idx:02d}",
            font=("Bebas Neue", 13, "bold"),
            text_color=PALETTE["accent"],
            width=30,
        ).pack(side="left", padx=(0, 10))
 
        title_row = ctk.CTkFrame(top, fg_color="transparent")
        title_row.pack(side="left", fill="x", expand=True)
 
        ctk.CTkLabel(
            title_row,
            text=self.job.title or "—",
            font=("Bebas Neue", 13, "bold"),
            text_color=PALETTE["text"],
            anchor="w",
        ).pack(side="left")
 
        if self.job.easy_apply:
            ctk.CTkLabel(
                title_row, text="⚡ Easy Apply",
                font=FONT_SMALL,
                text_color=PALETTE["accent2"],
                fg_color=PALETTE["tag_bg"],
                corner_radius=4,
                padx=6, pady=2,
            ).pack(side="left", padx=(10, 0))
 
        salary = getattr(self.job, "salary", None)
        if salary:
            ctk.CTkLabel(
                title_row, text=f"💰 {salary}",
                font=FONT_SMALL,
                text_color=PALETTE["accent2"],
                fg_color=PALETTE["tag_bg"],
                corner_radius=4,
                padx=6, pady=2,
            ).pack(side="left", padx=(8, 0))
 
        meta = ctk.CTkFrame(self, fg_color="transparent")
        meta.pack(fill="x", padx=54, pady=(0, 6))
 
        def _meta(text, color=PALETTE["muted"]):
            ctk.CTkLabel(
                meta, text=text, font=FONT_SMALL,
                text_color=color, anchor="w",
            ).pack(side="left", padx=(0, 16))
 
        _meta(f"🏢 {self.job.company or 'N/A'}", PALETTE["text"])
        _meta(f"📍 {self.job.location or 'N/A'}")
        _meta(f"📅 {self.job.date_posted or 'N/A'}")
        if self.job.num_applicants:
            _meta(f"👥 {self.job.num_applicants} applicants")
 
        if self.job.description:
            lines = [l.strip() for l in self.job.description.splitlines() if l.strip()][:2]
            preview = " • ".join(lines)[:180] + ("…" if len(" • ".join(lines)) > 180 else "")
            ctk.CTkLabel(
                self, text=preview,
                font=FONT_SMALL,
                text_color=PALETTE["muted"],
                anchor="w", wraplength=720,
            ).pack(fill="x", padx=54, pady=(0, 8))
 
        ctk.CTkLabel(
            self, text="click to view details →",
            font=("Bebas Neue", 9),
            text_color=PALETTE["border"],
            anchor="e",
        ).pack(fill="x", padx=14, pady=(0, 6))
 
 
# ─────────────────────────────────────────────
#  Mouse-wheel binding helpers
# ─────────────────────────────────────────────
 
import platform as _platform
_SYS = _platform.system()
 
 
def _make_scroll_handler(canvas):
    """Return a scroll callback bound to a specific canvas."""
    def _scroll(event):
        if _SYS == "Windows":
            canvas.yview_scroll(int(-1 * event.delta / 120), "units")
        elif _SYS == "Darwin":
            canvas.yview_scroll(int(-1 * event.delta), "units")
        else:
            canvas.yview_scroll(-1 if event.num == 4 else 1, "units")
    return _scroll
 
 
def _bind_tree(widget, handler):
    """Recursively bind scroll events on widget and all descendants.
    Deliberately only touches MouseWheel / Button-4 / Button-5 so it
    never clobbers or duplicates <Button-1> click handlers on cards."""
    if _SYS in ("Windows", "Darwin"):
        widget.bind("<MouseWheel>", handler, add="+")
    else:
        widget.bind("<Button-4>", handler, add="+")
        widget.bind("<Button-5>", handler, add="+")
    for child in widget.winfo_children():
        _bind_tree(child, handler)
 
 
def _attach_scroll(scrollable_frame: ctk.CTkScrollableFrame):
    """
    Bind mouse-wheel on the CTkScrollableFrame and every widget inside it
    so scrolling works regardless of which child the cursor is over.
    Safe to call multiple times — existing bindings use add='+'.
    """
    canvas = scrollable_frame._parent_canvas
    handler = _make_scroll_handler(canvas)
    _bind_tree(scrollable_frame, handler)
 
 
# Legacy alias — just call _attach_scroll directly
def _bind_mousewheel(widget):
    if hasattr(widget, "_parent_canvas"):
        _attach_scroll(widget)
 
 
# ─────────────────────────────────────────────
#  Main application window
# ─────────────────────────────────────────────
 
class LinkedInScraperApp(ctk.CTk):
 
    def __init__(self):
        super().__init__()
        self.title("LinkedIn Job Scraper")
        self.geometry("1160x860")
        self.minsize(960, 680)
        self.configure(fg_color=PALETTE["bg"])
 
        self._jobs: list[JobListing] = []
        self._running = False
        self._active_palette = PALETTE_DARK_KHAKI_ORANGE
 
        self._build_ui()
 
    # ─────────────────────────────────────────
    #  UI construction
    # ─────────────────────────────────────────
 
    def _build_ui(self):
        # ── Header ────────────────────────────
        header = ctk.CTkFrame(self, fg_color=PALETTE["panel"], corner_radius=0, height=54)
        header.pack(fill="x", side="top")
        header.pack_propagate(False)
 
        ctk.CTkLabel(
            header,
            text="◈  LINKEDIN  JOB  SCRAPER",
            font=FONT_TITLE,
            text_color=PALETTE["accent"],
        ).pack(side="left", padx=22, pady=10)
 
        self._status_label = ctk.CTkLabel(
            header, text="idle",
            font=FONT_MONO,
            text_color=PALETTE["muted"],
        )
        self._status_label.pack(side="right", padx=22)
 
        # ── Body ──────────────────────────────
        self._body = ctk.CTkFrame(self, fg_color="transparent")
        self._body.pack(fill="both", expand=True)
 
        self._build_left_panel(self._body)
        self._build_right_panel(self._body)
 
    def _build_left_panel(self, parent):
        left = ctk.CTkFrame(
            parent, fg_color=PALETTE["panel"],
            corner_radius=0, width=310,
        )
        left.pack(side="left", fill="y")
        left.pack_propagate(False)
 
        self._left_scroll = ctk.CTkScrollableFrame(
            left, fg_color="transparent",
            scrollbar_button_color=PALETTE["border"],
            scrollbar_button_hover_color=PALETTE["accent"],
        )
        self._left_scroll.pack(fill="both", expand=True, padx=16, pady=16)
        self.after(200, lambda: _attach_scroll(self._left_scroll))
 
        s = self._left_scroll
 
        # ── Theme switcher ────────────────────
        SectionLabel(s, "Color Theme").pack(anchor="w", pady=(0, 6))
        swatch_row = ctk.CTkFrame(s, fg_color="transparent")
        swatch_row.pack(anchor="w", pady=(0, 12))
 
        self._swatches: list[ThemeSwatch] = []
        for pal in ALL_PALETTES:
            sw = ThemeSwatch(swatch_row, pal, on_select=self._apply_palette)
            sw.pack(side="left", padx=(0, 6))
            self._swatches.append(sw)
 
        # Mark the active one
        self._swatches[0].set_active(True)
 
        # ── Keywords ──────────────────────────
        _divider(s)
        SectionLabel(s, "Keywords").pack(anchor="w", pady=(0, 4))
        self._kw = ctk.CTkEntry(
            s, placeholder_text="Python Developer, Data Scientist…",
            font=FONT_BODY, height=36,
            fg_color=PALETTE["card"], border_color=PALETTE["border"],
            text_color=PALETTE["text"],
        )
        self._kw.pack(fill="x", pady=(0, 12))
 
        # ── Location ──────────────────────────
        SectionLabel(s, "Location").pack(anchor="w", pady=(0, 4))
        self._loc = ctk.CTkEntry(
            s, placeholder_text="London, Remote, Worldwide…",
            font=FONT_BODY, height=36,
            fg_color=PALETTE["card"], border_color=PALETTE["border"],
            text_color=PALETTE["text"],
        )
        self._loc.pack(fill="x", pady=(0, 12))
 
        # ── Date filter ───────────────────────
        SectionLabel(s, "Date Posted").pack(anchor="w", pady=(0, 4))
        self._date_var = ctk.StringVar(value="Any time")
        self._date_menu = ctk.CTkOptionMenu(
            s, variable=self._date_var,
            values=list(DATE_OPTIONS.keys()),
            font=FONT_BODY, height=34,
            fg_color=PALETTE["card"], button_color=PALETTE["border"],
            button_hover_color=PALETTE["accent"],
            dropdown_fg_color=PALETTE["card"],
            text_color=PALETTE["text"],
        )
        self._date_menu.pack(fill="x", pady=(0, 6))
 
        self._custom_secs_var = ctk.BooleanVar()
        ctk.CTkCheckBox(
            s, text="Custom time window (seconds)",
            variable=self._custom_secs_var,
            font=FONT_SMALL, text_color=PALETTE["muted"],
            fg_color=PALETTE["accent"], border_color=PALETTE["border"],
            hover_color=self._accent_hover(),
            command=self._toggle_custom_secs,
        ).pack(anchor="w", pady=(0, 4))
 
        # Wrapper is always packed; the entry inside is shown/hidden
        self._custom_secs_wrapper = ctk.CTkFrame(s, fg_color="transparent", height=0)
        self._custom_secs_wrapper.pack(fill="x")
        self._custom_secs_entry = ctk.CTkEntry(
            self._custom_secs_wrapper, placeholder_text="e.g. 3600",
            font=FONT_BODY, height=32,
            fg_color=PALETTE["card"], border_color=PALETTE["border"],
            text_color=PALETTE["text"],
        )
        # entry is hidden by default — shown when checkbox is ticked
 
        # ── Sort ──────────────────────────────
        _divider(s)
        SectionLabel(s, "Sort By").pack(anchor="w", pady=(0, 4))
        self._sort_var = ctk.StringVar(value="Most recent")
        ctk.CTkOptionMenu(
            s, variable=self._sort_var,
            values=list(SORT_OPTIONS.keys()),
            font=FONT_BODY, height=34,
            fg_color=PALETTE["card"], button_color=PALETTE["border"],
            button_hover_color=PALETTE["accent"],
            dropdown_fg_color=PALETTE["card"],
            text_color=PALETTE["text"],
        ).pack(fill="x", pady=(0, 12))
 
        # ── Max results ───────────────────────
        SectionLabel(s, "Max Results").pack(anchor="w", pady=(0, 4))
        self._max_results = ctk.CTkEntry(
            s, placeholder_text="25",
            font=FONT_BODY, height=32,
            fg_color=PALETTE["card"], border_color=PALETTE["border"],
            text_color=PALETTE["text"],
        )
        self._max_results.insert(0, "25")
        self._max_results.pack(fill="x", pady=(0, 12))
 
        # ── Options ───────────────────────────
        self._fetch_desc_var = ctk.BooleanVar(value=True)
        ctk.CTkCheckBox(
            s, text="Fetch full descriptions",
            variable=self._fetch_desc_var,
            font=FONT_SMALL, text_color=PALETTE["text"],
            fg_color=PALETTE["accent"], border_color=PALETTE["border"],
            hover_color=self._accent_hover(),
        ).pack(anchor="w", pady=(0, 6))
 
        self._headless_var = ctk.BooleanVar(value=True)
        ctk.CTkCheckBox(
            s, text="Headless browser",
            variable=self._headless_var,
            font=FONT_SMALL, text_color=PALETTE["text"],
            fg_color=PALETTE["accent"], border_color=PALETTE["border"],
            hover_color=self._accent_hover(),
        ).pack(anchor="w", pady=(0, 12))
 
        # ══════════════════════════════════════
        #  ADVANCED FILTERS  (collapsible)
        # ══════════════════════════════════════
        _divider(s)
        self._adv_section = CollapsibleSection(s, "Advanced Filters", initially_open=False)
        self._adv_section.pack(fill="x", pady=(0, 8))
 
        adv = self._adv_section.content  # inner frame
 
        # ── Experience Level ──────────────────
        SectionLabel(adv, "Experience Level").pack(anchor="w", pady=(4, 4))
        self._exp_group = MultiCheckGroup(adv, EXPERIENCE_OPTIONS)
        self._exp_group.pack(fill="x", pady=(0, 10))
 
        # ── Job Type ──────────────────────────
        ctk.CTkFrame(adv, fg_color=PALETTE["border"], height=1).pack(fill="x", pady=6)
        SectionLabel(adv, "Job Type").pack(anchor="w", pady=(0, 4))
        self._jtype_group = MultiCheckGroup(adv, JOB_TYPE_OPTIONS)
        self._jtype_group.pack(fill="x", pady=(0, 10))
 
        # ── Work Type ─────────────────────────
        ctk.CTkFrame(adv, fg_color=PALETTE["border"], height=1).pack(fill="x", pady=6)
        SectionLabel(adv, "Work Type").pack(anchor="w", pady=(0, 4))
        self._wtype_group = MultiCheckGroup(adv, WORK_TYPE_OPTIONS)
        self._wtype_group.pack(fill="x", pady=(0, 10))
 
        # ── Toggles ───────────────────────────
        ctk.CTkFrame(adv, fg_color=PALETTE["border"], height=1).pack(fill="x", pady=6)
        self._easy_apply_var = ctk.BooleanVar()
        ctk.CTkCheckBox(
            adv, text="⚡  Easy Apply only",
            variable=self._easy_apply_var,
            font=FONT_BODY, text_color=PALETTE["text"],
            fg_color=PALETTE["accent"], border_color=PALETTE["border"],
            hover_color=self._accent_hover(),
        ).pack(anchor="w", pady=(0, 8))
 
        self._active_var = ctk.BooleanVar()
        ctk.CTkCheckBox(
            adv, text="🔥  Actively Hiring only",
            variable=self._active_var,
            font=FONT_BODY, text_color=PALETTE["text"],
            fg_color=PALETTE["accent"], border_color=PALETTE["border"],
            hover_color=self._accent_hover(),
        ).pack(anchor="w", pady=(0, 10))
 
        # ── Search Radius ─────────────────────
        ctk.CTkFrame(adv, fg_color=PALETTE["border"], height=1).pack(fill="x", pady=6)
        SectionLabel(adv, "Search Radius (miles)").pack(anchor="w", pady=(0, 4))
        self._distance = ctk.CTkEntry(
            adv, placeholder_text="25, 50… (optional)",
            font=FONT_BODY, height=32,
            fg_color=PALETTE["card"], border_color=PALETTE["border"],
            text_color=PALETTE["text"],
        )
        self._distance.pack(fill="x", pady=(0, 10))
 
        # ── Run button ────────────────────────
        _divider(s)
        self._run_btn = ctk.CTkButton(
            s,
            text="▶  SEARCH",
            font=("Bebas Neue", 13, "bold"),
            height=42,
            fg_color=PALETTE["accent"],
            hover_color=self._accent_hover(),
            text_color="#ffffff",
            corner_radius=6,
            command=self._start_search,
        )
        self._run_btn.pack(fill="x", pady=(0, 8))
 
        # ── Export buttons ────────────────────
        export_row = ctk.CTkFrame(s, fg_color="transparent")
        export_row.pack(fill="x", pady=(0, 4))
 
        self._save_json_btn = ctk.CTkButton(
            export_row,
            text="💾 JSON",
            font=FONT_SMALL,
            height=34,
            fg_color=PALETTE["card"],
            hover_color=PALETTE["card_hover"],
            border_width=1,
            border_color=PALETTE["border"],
            text_color=PALETTE["muted"],
            corner_radius=6,
            command=self._export_json,
            state="disabled",
        )
        self._save_json_btn.pack(side="left", fill="x", expand=True, padx=(0, 4))
 
        self._save_excel_btn = ctk.CTkButton(
            export_row,
            text="📊 Excel",
            font=FONT_SMALL,
            height=34,
            fg_color=PALETTE["card"],
            hover_color=PALETTE["card_hover"],
            border_width=1,
            border_color=PALETTE["border"],
            text_color=PALETTE["muted"],
            corner_radius=6,
            command=self._export_excel,
            state="disabled",
        )
        self._save_excel_btn.pack(side="left", fill="x", expand=True)
 
    def _build_right_panel(self, parent):
        self._right = ctk.CTkFrame(parent, fg_color="transparent")
        self._right.pack(side="left", fill="both", expand=True, padx=12, pady=12)
 
        self._progress = ctk.CTkProgressBar(
            self._right, mode="indeterminate",
            fg_color=PALETTE["border"],
            progress_color=PALETTE["accent"],
            height=3,
        )
 
        self._results_frame = ctk.CTkScrollableFrame(
            self._right, fg_color="transparent",
            scrollbar_button_color=PALETTE["border"],
            scrollbar_button_hover_color=PALETTE["accent"],
        )
        self._results_frame.pack(fill="both", expand=True)
        self.after(200, lambda: _attach_scroll(self._results_frame))
 
        self._empty_label = ctk.CTkLabel(
            self._results_frame,
            text="Enter search parameters and click  ▶ SEARCH",
            font=("Bebas Neue", 13),
            text_color=PALETTE["muted"],
        )
        self._empty_label.pack(expand=True, pady=140)
 
    # ─────────────────────────────────────────
    #  Theme switching
    # ─────────────────────────────────────────
 
    def _apply_palette(self, palette: dict):
        """Switch the active palette and rebuild the entire UI."""
        global PALETTE
        PALETTE.update(palette)
        self._active_palette = palette
 
        mode = palette.get("_mode", "dark")
        ctk.set_appearance_mode(mode)
 
        # Update swatch borders
        for sw in self._swatches:
            sw.set_active(sw._palette is palette)
 
        # Rebuild the whole window
        self.configure(fg_color=PALETTE["bg"])
        for widget in self.winfo_children():
            widget.destroy()
        self._jobs = []
        self._running = False
        self._build_ui()
 
    def _accent_hover(self):
        acc = PALETTE["accent"]
        r = max(0, int(acc[1:3], 16) - 30)
        g = max(0, int(acc[3:5], 16) - 30)
        b = max(0, int(acc[5:7], 16) - 30)
        return f"#{r:02x}{g:02x}{b:02x}"
 
    # ─────────────────────────────────────────
    #  Helpers
    # ─────────────────────────────────────────
 
    def _toggle_custom_secs(self):
        if self._custom_secs_var.get():
            self._date_menu.configure(state="disabled")
            self._custom_secs_wrapper.configure(height=40)
            self._custom_secs_entry.pack(fill="x", padx=0, pady=(0, 4))
            self._custom_secs_entry.focus()
        else:
            self._date_menu.configure(state="normal")
            self._custom_secs_entry.pack_forget()
            self._custom_secs_wrapper.configure(height=0)
 
    def _set_status(self, text: str, color: str = None):
        if color is None:
            color = PALETTE["muted"]
        self._status_label.configure(text=text, text_color=color)
 
    def _clear_results(self):
        for widget in self._results_frame.winfo_children():
            widget.destroy()
        # Reset live-counter refs so stale widgets are never accessed
        self._scrape_counter_lbl = None
        self._scrape_phase_lbl = None
        self._scrape_counter_frame = None
 
    # ─────────────────────────────────────────
    #  Detail view navigation
    # ─────────────────────────────────────────
 
    def _open_detail(self, job: JobListing):
        self._results_frame.pack_forget()
        if hasattr(self, "_progress"):
            self._progress.pack_forget()
 
        self._detail_view = JobDetailView(
            self._right, job,
            on_back=self._close_detail,
        )
        self._detail_view.pack(fill="both", expand=True)
        # Bind scroll directly on the named scrollable body — no index guessing
        self.after(100, lambda: _attach_scroll(self._detail_view._body))
 
    def _close_detail(self):
        if hasattr(self, "_detail_view"):
            self._detail_view.destroy()
            del self._detail_view
        self._results_frame.pack(fill="both", expand=True)
 
    # ─────────────────────────────────────────
    #  Search
    # ─────────────────────────────────────────
 
    def _start_search(self):
        keywords = self._kw.get().strip()
        if not keywords:
            messagebox.showwarning("Missing keywords", "Please enter job title or keywords.")
            return
        if self._running:
            return
 
        self._running = True
        self._run_btn.configure(state="disabled", text="⏳ Searching…")
        self._save_json_btn.configure(state="disabled")
        self._save_excel_btn.configure(state="disabled")
        self._clear_results()
 
        config = self._build_config()
        max_results = int(self._max_results.get() or 25)
        fetch_desc = self._fetch_desc_var.get()
        headless = self._headless_var.get()
 
        # ── Live progress UI ──────────────────
        self._progress.pack(fill="x", pady=(0, 6))
        self._progress.start()
 
        # Counter bar  (e.g.  "scraped  7 / 25")
        self._scrape_counter_frame = ctk.CTkFrame(
            self._results_frame,
            fg_color=PALETTE["panel"],
            corner_radius=6,
            border_width=1,
            border_color=PALETTE["border"],
        )
        self._scrape_counter_frame.pack(fill="x", pady=(0, 8))
 
        self._scrape_counter_lbl = ctk.CTkLabel(
            self._scrape_counter_frame,
            text=f"  scraped  0 / {max_results}",
            font=("Bebas Neue", 11, "bold"),
            text_color=PALETTE["accent"],
            anchor="w",
        )
        self._scrape_counter_lbl.pack(side="left", padx=12, pady=8)
 
        self._scrape_phase_lbl = ctk.CTkLabel(
            self._scrape_counter_frame,
            text="initialising…",
            font=FONT_MONO,
            text_color=PALETTE["muted"],
            anchor="e",
        )
        self._scrape_phase_lbl.pack(side="right", padx=12, pady=8)
 
        self._set_status("searching…", PALETTE["accent"])
 
        # Shared live job list written by scraper thread, read by UI thread
        self._live_jobs: list[JobListing] = []
        self._live_max = max_results
 
        def _on_job(job: JobListing):
            """Called from scraper thread each time one job is ready."""
            self._live_jobs.append(job)
            n = len(self._live_jobs)
            self.after(0, lambda job=job, n=n: self._append_live_card(job, n))
 
        def _on_phase(phase: str):
            """Called from scraper thread when the phase changes."""
            self.after(0, lambda p=phase: self._scrape_phase_lbl.configure(text=p))
 
        def run():
            try:
                url = build_url(config)
                scraper = LinkedInScraper(
                    headless=headless,
                    fetch_descriptions=fetch_desc,
                )
                # Pass callbacks if the scraper supports them; fall back gracefully
                try:
                    jobs = asyncio.run(
                        scraper.scrape(
                            url,
                            max_results=max_results,
                            on_job=_on_job,
                            on_phase=_on_phase,
                        )
                    )
                except TypeError:
                    # Scraper does not accept callbacks — run normally, stream at end
                    _on_phase("scraping…")
                    jobs = asyncio.run(scraper.scrape(url, max_results=max_results))
                    for job in jobs:
                        _on_job(job)
 
                self.after(0, lambda: self._finalise_results(jobs, config))
            except Exception as e:
                self.after(0, lambda: self._on_error(str(e)))
 
        threading.Thread(target=run, daemon=True).start()
 
    def _build_config(self) -> SearchConfig:
        custom_secs = None
        date_filter = "any"
        if self._custom_secs_var.get():
            raw = self._custom_secs_entry.get().strip()
            custom_secs = int(raw) if raw.isdigit() else 1800
        else:
            date_filter = DATE_OPTIONS.get(self._date_var.get(), "any")
 
        dist_raw = self._distance.get().strip()
        distance = int(dist_raw) if dist_raw.isdigit() else None
 
        return SearchConfig(
            keywords=self._kw.get().strip(),
            location=self._loc.get().strip() or "Worldwide",
            date_filter=date_filter,
            custom_seconds=custom_secs,
            experience=self._exp_group.selected(),
            job_type=self._jtype_group.selected(),
            work_type=self._wtype_group.selected(),
            easy_apply=self._easy_apply_var.get(),
            actively_hiring=self._active_var.get(),
            sort_by=SORT_OPTIONS.get(self._sort_var.get(), "recent"),
            distance=distance,
        )
 
    def _append_live_card(self, job: JobListing, n: int):
        """Add one card to the results panel as it arrives from the scraper thread."""
        max_label = self._live_max if hasattr(self, "_live_max") else "?"
 
        # Update counter label — guard against stale widget after a re-search
        try:
            lbl = self._scrape_counter_lbl
            if lbl.winfo_exists():
                lbl.configure(text=f"  scraped  {n} / {max_label}")
        except Exception:
            pass
 
        self._set_status(f"scraped {n}…", PALETTE["accent"])
 
        card = JobCard(self._results_frame, job, n, on_click=self._open_detail)
        card.pack(fill="x", pady=(0, 6))
        _attach_scroll(self._results_frame)
 
    def _finalise_results(self, jobs: list[JobListing], config: SearchConfig):
        """Called once scraping is fully complete."""
        self._running = False
        self._progress.stop()
        self._progress.pack_forget()
        self._run_btn.configure(state="normal", text="▶  SEARCH")
        self._jobs = jobs
 
        if not jobs:
            # Remove counter bar and show empty message
            if hasattr(self, "_scrape_counter_frame"):
                self._scrape_counter_frame.destroy()
            ctk.CTkLabel(
                self._results_frame,
                text="No results found. Try adjusting your filters.",
                font=FONT_BODY, text_color=PALETTE["muted"],
            ).pack(pady=80)
            self._set_status("0 results", PALETTE["warn"])
            return
 
        # Update counter bar to final summary
        if hasattr(self, "_scrape_counter_lbl"):
            self._scrape_counter_lbl.configure(
                text=f"  {len(jobs)} results  ·  {config.keywords}  ·  {config.location}",
                text_color=PALETTE["accent2"],
            )
        if hasattr(self, "_scrape_phase_lbl"):
            self._scrape_phase_lbl.configure(text="done ✓", text_color=PALETTE["accent2"])
 
        self._set_status(f"{len(jobs)} jobs found", PALETTE["accent2"])
        self.after(100, lambda: _attach_scroll(self._results_frame))
 
        self._save_json_btn.configure(state="normal", text_color=PALETTE["text"])
        self._save_excel_btn.configure(state="normal", text_color=PALETTE["text"])
 
    def _on_error(self, msg: str):
        self._running = False
        self._progress.stop()
        self._progress.pack_forget()
        self._run_btn.configure(state="normal", text="▶  SEARCH")
        if hasattr(self, "_scrape_counter_frame"):
            try:
                self._scrape_phase_lbl.configure(text="error ✗", text_color=PALETTE["warn"])
            except Exception:
                pass
        self._set_status("error", PALETTE["warn"])
        messagebox.showerror("Scraper error", msg)
 
    # ─────────────────────────────────────────
    #  Export
    # ─────────────────────────────────────────
 
    def _export_json(self):
        if not self._jobs:
            return
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"linkedin_jobs_{timestamp}.json"
        with open(filename, "w", encoding="utf-8") as f:
            json.dump(
                [j.to_dict() for j in self._jobs],
                f, indent=2, ensure_ascii=False,
            )
        messagebox.showinfo("Exported", f"Saved {len(self._jobs)} jobs → {filename}")
 
    def _export_excel(self):
        if not self._jobs:
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            messagebox.showerror(
                "Missing dependency",
                "openpyxl is required for Excel export.\n\nRun:  pip install openpyxl",
            )
            return
 
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"linkedin_jobs_{timestamp}.xlsx"
 
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "LinkedIn Jobs"
 
        p = PALETTE
        hdr_hex  = p["panel"].lstrip("#")
        body_hex = p["card"].lstrip("#")
        alt_hex  = p["card_hover"].lstrip("#")
        acc_hex  = p["accent"].lstrip("#")
 
        HDR_FILL  = PatternFill("solid", fgColor=hdr_hex)
        HDR_FONT  = Font(name="Bebas Neue", bold=True, color=acc_hex, size=11)
        BODY_FILL = PatternFill("solid", fgColor=body_hex)
        ALT_FILL  = PatternFill("solid", fgColor=alt_hex)
        BODY_FONT = Font(name="Bebas Neue", color=p["text"].lstrip("#"), size=10)
        LINK_FONT = Font(name="Bebas Neue", color=acc_hex, size=10, underline="single")
        BORDER    = Border(
            bottom=Side(style="thin", color=p["border"].lstrip("#")),
            right=Side(style="thin", color=p["border"].lstrip("#")),
        )
 
        headers = [
            "#", "Job Title", "Company", "Location",
            "Date Posted", "Applicants", "Easy Apply",
            "Salary", "Work Type", "URL",
        ]
 
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = HDR_FONT
            cell.fill = HDR_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = BORDER
 
        ws.row_dimensions[1].height = 22
 
        for row_idx, job in enumerate(self._jobs, start=2):
            salary = getattr(job, "salary", "") or ""
            work_type = getattr(job, "work_type", "") or ""
            easy = "Yes" if job.easy_apply else "No"
 
            values = [
                row_idx - 1,
                job.title or "",
                job.company or "",
                job.location or "",
                job.date_posted or "",
                job.num_applicants or "",
                easy,
                salary,
                work_type,
                job.url or "",
            ]
 
            fill = BODY_FILL if row_idx % 2 == 0 else ALT_FILL
 
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.fill = fill
                cell.border = BORDER
                cell.alignment = Alignment(vertical="center", wrap_text=False)
                if col == len(headers) and job.url:
                    cell.font = LINK_FONT
                else:
                    cell.font = BODY_FONT
 
            ws.row_dimensions[row_idx].height = 18
 
        col_widths = [5, 36, 24, 22, 14, 12, 12, 16, 12, 50]
        for col, width in enumerate(col_widths, 1):
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col)
            ].width = width
 
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(headers))}1"
 
        wb.save(filename)
        messagebox.showinfo("Exported", f"Saved {len(self._jobs)} jobs → {filename}")
 
 
# ─────────────────────────────────────────────
#  Entry point
# ─────────────────────────────────────────────
 
def main():
    app = LinkedInScraperApp()
    app.mainloop()
 
 
if __name__ == "__main__":
    main()